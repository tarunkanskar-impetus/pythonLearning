from selenium import webdriver
from openpyxl import load_workbook
from TestPackage.Login import element



# Read Data from Excel File
wb = load_workbook(filename='D:/PythonLearning/PythonLearning.xlsx')
ws = wb.active
username = ws['A2'].value
password = ws['B2'].value



driver = webdriver.Chrome(executable_path='D:/Setups/chromedriver_win32/chromedriver.exe')
driver.get("http://ec2-18-223-116-199.us-east-2.compute.amazonaws.com/")
driver.maximize_window()
driver.implicitly_wait(30)


#Fill Login Details
driver.find_element_by_id('username').send_keys(username)
driver.find_element_by_id('password').send_keys(password)
driver.find_element_by_name('submit').click()
driver.set_page_load_timeout(10)

#Click Edit Profile
driver.find_element_by_link_text('Edit Profile').click()

#Update First Name
driver.find_element_by_id('id_first_name').clear()
driver.find_element_by_id('id_first_name').send_keys('Tarun2')
driver.find_element_by_name('submit').click()

driver.find_element_by_link_text('Dashboard').click()
driver.find_element_by_link_text('Edit Profile').click()
element = driver.find_element_by_id('id_first_name')
print(element.text)

