from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from openpyxl import load_workbook



# Read Data from Excel File
wb = load_workbook(filename='D:/PythonLearning/PythonLearning.xlsx')
ws = wb.active
username = ws['A2'].value
password = ws['B2'].value

# Open Browser
driver = webdriver.Chrome(executable_path='D:/Setups/chromedriver_win32/chromedriver.exe')
driver.get("http://ec2-18-223-116-199.us-east-2.compute.amazonaws.com/")
driver.implicitly_wait(30)
driver.maximize_window()

#Fill Login Details
driver.find_element_by_id('username').send_keys(username)
driver.find_element_by_id('password').send_keys(password)
driver.find_element_by_name('submit').click()
driver.set_page_load_timeout(10)


#Take Screen Shot
try:
    element = WebDriverWait(driver, 10).until(
        EC.presence_of_all_elements_located((By.CLASS_NAME,"table-bordered"))
        
    )
finally:
    driver.get_screenshot_as_file('D:/PythonWorkSpace/CurrentWeather.png') 
    driver.quit()
    
    
    #LogOut