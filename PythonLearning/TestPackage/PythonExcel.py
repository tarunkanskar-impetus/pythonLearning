from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook(filename='D:/PythonLearning/PythonLearning.xlsx')
ws = wb.active

print("highest row:" + str(ws.get_highest_row()))
print('str')
b1= ws['A1']
print("b1 --> ", b1.value)